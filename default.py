# -*- coding: utf-8 -*-
# -------------------------------------------------------------------------
# This is a sample controller
# this file is released under public domain and you can use without limitations
# -------------------------------------------------------------------------
import os
from urllib.parse import quote
import openpyxl

#envs=[(1,'DEV'),(2,'INT'),(3,'PrePROD'),(4,'PROD')]
envs=[(1,'Development'),(2,'Integration'),(3,'Qualification'),(4,'Preproduction'),(5,'Production')]
port=8608  # le port sur lequel le service plantu ecoute sur l'hote (et non pas sur le container)

envs_d=dict(envs)

def _init_log():
    import os,logging,logging.handlers,time 
    logger = logging.getLogger(request.application) 
    logger.setLevel(logging.INFO)
    #py2to3
    # ajout du param encoding
    handler = logging.handlers.RotatingFileHandler(os.path.join(
        request.folder,'logs','applog2.log'),'a',1024*1024,1,encoding='utf-8')
    handler.setLevel(logging.INFO) #or DEBUG / INFO
    handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s %(filename)s %(lineno)d %(funcName)s(): %(message)s')) 
    logger.addHandler(handler) 
    return logger

app_logging = cache.ram('app_wide_log',lambda:_init_log(),time_expire=None)

# ---- example index page ----
def index():
    response.flash = T("Hello World")
    return dict(message=T('Welcome to web2py!'))

# ---- Action for login/register/etc (required for auth) -----
def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/bulk_register
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    also notice there is http://..../[app]/appadmin/manage/auth to allow administrator to manage users
    """
    return dict(form=auth())


def import_pad(): 
    from gluon.sqlhtml import form_factory

    fi0=SQLField('import_xls','upload',requires=IS_NOT_EMPTY("Please provide a valid PAD file"))
    fi1=Field('envi',requires=IS_IN_SET(envs),default=1,widget=SQLFORM.widgets.radio.widget,label='Environment')
    
    fields=[fi0,fi1]

    form = form_factory(*fields, 
        submit_button='Generate a UML schema',
               formstyle='table2cols'
                )

    #form=form_factory(,)
    if form.accepts(request.vars,session): 
        # champs correctement remplis

        #request.flash='Received: %s'%request.vars.import_xls
        #app_logging.info(request.vars.envi)

        #print (request.vars.import_xls.filename)
        fobj=request.vars.import_xls
        # ligne suivante indispensable et totalement inexplicable :)
        str(fobj)

        fissier=fobj.file.read()
        taille_fic=len(fissier)

        response.flash='Uploaded PAD : %s, size is %s bytes'%(fobj.filename, taille_fic)

        PADxl_path='applications/%s/static/upload/%s' % (request.application,request.vars.import_xls.filename)
        environment=request.vars.envi
        
        # ecrit le fichier uploadé sur le filesystem
        ficout=open(file=PADxl_path,mode='wb')
        ficout.write(fissier)
        ficout.close()

        plantu=create_plantu(PADxl_path,environment)
        #app_logging.info("Voila le resulat")
        #app_logging.info(plantu)

        # on doit le convertir en "percent encoding" pour qu'il passe en parametre dans l'url de 3 km de long
        plantu_clean=quote(plantu, safe='')

        # génère l'image en redirigeant sur le service plantu-service
        host=request.env.http_host
        #app_logging.info(request.env)
        # meme hote que l'application web mais service plantu (svg) grace au reverse-proxy
        URL="https://%s/svg/%s" % (host,plantu_clean)

        #app_logging.info("URL:%s" % URL)

        redirect(URL)

    return dict(form=form)

# recherche le tableau dans un onglet, et en donne les dimmensions
def read_tab(wbook, tab_name, keyword_begin, keyword_end):
    # variabled e resultat
    begin_row = 0
    last_row = 0
    # recherche de l onglet
    sheet = wbook.get_sheet_by_name(tab_name)
    # parcourt de la feuille
    rows = sheet.rows
    n_row = 1
    for row in rows:
        n_cell = 1
        for cell in row:
            if cell.data_type != "n":
                if cell.value == keyword_begin:
                    begin_row = n_row
                    print("begin ={0}".format(cell.coordinate))
                if begin_row > 0 and cell.value == keyword_end:
                    print("last  ={0}".format(cell.coordinate))
                    last_row = n_row
            n_cell += 1
        n_row += 1
    # resultat
    return sheet, begin_row, last_row

# lit la description de l application
def header_pu(wbook, env):
    sheetAppli = wbook.get_sheet_by_name("Application Summary")
    appName = sheetAppli.cell(4, 3).value
    cwCode = sheetAppli.cell(5, 3).value

    # Plant UML : header
    puml = "@startuml\n"
    # puml += "!include <cloudinsight/sqlserver>\n"
    puml += "!define DateGen %date[yyyy.MM.dd 'at' HH:MM]%\n"

    # Plant UML : contecxte application
    puml += "center header ARC@CA-GIP implementation diagram\ntitle\n"
    puml += "<i>Application</i> : <back:LimeGreen><b>%s</b></back>\n" % appName
    puml += "<i>Environment</i> : <u>%s</u>\n" % env
    if cwCode is not None and cwCode != "":
        puml += "<i>Code Appli</i> : <b>%s</b>\n" % cwCode
    puml += "end title\ncenter footer generated DateGen\n"
    puml += "\n"

    return puml

# lit l onglet users, select env, ajoute a item
def user_pu(wbook, env):
    # recherche du tableau des users
    my_sheet, begin_row, last_row = read_tab (wbook, "Users & IAM", "Authentication", "Citrix")

    puml, nb, items = "\n", 0, []
    for x in range(begin_row + 2, last_row - 1):
        envi = my_sheet.cell(x, 2).value
        name = my_sheet.cell(x, 4).value
        loca = my_sheet.cell(x, 5).value
        if envi == env:
            puml += "cloud \"%s\" {\n" % loca
            puml += "actor \"%s\" as item%s\n}\n" % (name, str(nb))
            items.append(name)
            nb += 1
    # resultat
    return items, puml

# lit l onglet database, select env, ajoute a item
def data_pu(wbook, env, items):
    # recherche du tableau des data
    my_sheet, begin_row, last_row = read_tab(wbook, "Data Storage & Security", "Database", "Database Volume")

    nb, names, descs, new_item = len(items), [], [], []
    for x in range(begin_row + 2, last_row - 1):
        envi = my_sheet.cell(x, 2).value
        clus = my_sheet.cell(x, 9).value
        inst = my_sheet.cell(x, 10).value
        if envi == env:
            desc = "database \"%s\" as item%s\n" % (inst, str(nb))
            # le cluster est-il deja dans la liste ?
            try:
                i = names.index(clus)
                # oui, on l a trouve, on les remplace
                desc += descs.pop(i)
                descs.append(desc)
                names.pop(i)
                names.append(clus)
            except ValueError:
                # non, on le rajoute
                names.append(clus)
                descs.append(desc)
            new_item.append(inst)
            nb += 1
    # resultat
    return new_item, names, descs

# lit l onglet filershare, select env, ajoute a item
def filer_pu(wbook, env, items):
    # recherche du tableau des fileshare
    my_sheet, begin_row, last_row = read_tab(wbook, "Data Storage & Security", "Fileshare", "Fileshare Volume")

    nb, ids, elements, new_item = len(items), [], [], []
    for x in range(begin_row + 2, last_row - 1):
        envi = my_sheet.cell(x, 2).value
        name = my_sheet.cell(x, 3).value
        host = my_sheet.cell(x, 5).value
        if envi == env:
            desc = "file \"%s\" as item%s\n" % (name, str(nb))
            # le host est-il deja dans la liste ?
            try:
                i = ids.index(host)
                # oui, on l a trouve, on les remplace
                desc += elements.pop(i)
                elements.append(desc)
                ids.pop(i)
                ids.append(host)
            except ValueError:
                # non, on le rajoute
                ids.append(host)
                elements.append(desc)
            new_item.append(name)
            nb += 1
    # resultat
    return new_item, ids, elements

# lit l onglet physical, select env, ajoute dans items
def physical_pu(wbook, env, items, ds_names, ds_descs, fi_names, fi_descs):
    # recherche du tableau des composants physiques
    my_sheet, begin_row, last_row = read_tab(wbook, "Physical", "Server ressources", "Network locations")

    # Plant UML : Component
    puml, nb, new_item = "\n", len(items), []
    for x in range(begin_row + 2, last_row - 2):
        envir = my_sheet.cell(x, 2).value
        statu = my_sheet.cell(x, 3).value
        name  = my_sheet.cell(x, 4).value
        funct = my_sheet.cell(x, 5).value
        s_RAM = my_sheet.cell(x, 10).value
        s_dsk = my_sheet.cell(x, 11).value
        local = my_sheet.cell(x, 12).value
        element = ""

        if statu != "Decommissionned" and envir == env:
            puml+="cloud \"%s\" {\n" % local
            # Database
            if funct == "Database":
                puml += "database \"%s\" as item%s\n" % (name, str(nb))
            # MQ server
            elif funct == "MQ server":
                puml += "queue \"%s\" as item%s\n" % (name, str(nb))
            # Component : recherche si present dans DataSource et Filer
            else:
                # recherche dans datasource
                try:
                    i = ds_names.index(name)
                    # oui, on l a trouve, on ajoute la description
                    element += "\n%s" % ds_descs[i]
                except ValueError:
                    # non, absent
                    element = element

                # recherche dans filer
                try:
                    i = fi_names.index(name)
                    # oui, on l a trouve, on ajoute la description
                    element += "\n%s" % fi_descs[i]
                except ValueError:
                    # non, absent
                    element = element

                if element == "":
                    # composant
                    puml += "component \"%s\\n%s\" as item%s\n" % (name, funct, str(nb))
                else :
                    # node avec sous composants
                    puml += "node \"%s (%s)\" as item%s {\n" % (name, funct, str(nb))
                    puml += "%s}\n" % element

            new_item.append(name)
            nb += 1
            # Note pour les New
            if statu == "New":
                puml += "note right of item%s\nRAM=%s\nDisk=%s\nend note\n" % (str(nb-1), s_RAM, s_dsk)

    return new_item, puml

# lit l onglet flux, select env, ajoute les nx composants
# @todo filterer sur environement
def flux_pu(wbook, env, items):
    #  recherche du tableau des data
    my_sheet, begin_row, last_row = read_tab(wbook, "Logical", "Flow", "End Flow")

    # Plant UML : LES FLUX
    puml, nb = "", len(items)
    for x in range(begin_row + 3, last_row - 1):
        source = str(my_sheet.cell(x, 6).value)
        dest = str(my_sheet.cell(x, 8).value)
        i_src = -1
        i_dest = -1
        # print("({0}) {1} --> {2}".format(x,source, dest))
        try:
            i_src = items.index(source)
        except ValueError:
            # print("external [" + str(x) + "] " + source)
            puml += "node \"%s\" as item%s\n" % (source, str(nb))
            items.append(source)
            i_src = nb
            nb += 1
        try:
            i_dest = items.index(dest)
        except ValueError:
            # print("external [" + str(x) + "] " + dest)
            puml += "node \"%s\" as item%s\n" % (dest, str(nb))
            items.append(dest)
            i_dest = nb
            nb += 1
        if i_src > -1 and i_dest > -1:
            num = str(my_sheet.cell(x, 2).value)
            prot = str(my_sheet.cell(x, 10).value)
            puml += "item%s --> item%s : %s:%s\n" % (str(i_src), str(i_dest), num, prot)

    return puml

def create_plantu(xls_file, envi):
    # genere le plantuml
    # ensemble de commandes qui modelisent le schéma
    # on stocke cette liste de commandes dans une variable
    # ce sera l'argument envoyé au webservice de génération d'image

    # Ouverture des fichiers
    wbook = openpyxl.load_workbook(xls_file, data_only=True)

    # le libelle de environement a partir du code dans le formulaire
    environment = envs_d[int(envi)]

    # le header PlantUML
    puml = header_pu(wbook ,environment)

    # la liste des items
    items = []
    # les acteurs
    items, p_user = user_pu(wbook, environment)
    puml += p_user

    # les datasources
    datasources, ds_names, ds_descs = data_pu(wbook, environment, items)
    items.extend(datasources)

    # les filers
    files, fi_names, fi_descs = filer_pu(wbook, environment, items)
    items.extend(files)

    # les acteurs composants physiques
    composants, p_compo = physical_pu(wbook, environment, items, ds_names, ds_descs, fi_names, fi_descs)
    puml += p_compo
    items.extend(composants)

    # les flux
    puml += flux_pu(wbook, environment, items)

    # fermeture fichier PlantUML
    puml += "@enduml\n"

    return puml
    